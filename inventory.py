
"""
Inventory Management Page with Excel Import (Performance Optimized v2)
Improved tabular alignment and database query optimization
File: inventory.py
"""

import customtkinter as ctk
from db import (
    add_product, get_all_products, update_product,
    delete_product, get_low_stock_products
)
import tkinter.filedialog as filedialog
from openpyxl import load_workbook
import threading
from functools import lru_cache

class Page(ctk.CTkFrame):
    def __init__(self, parent, go_back_callback):
        super().__init__(parent, fg_color="#f8f9fa")
        self.go_back = go_back_callback

        self.grid_rowconfigure(2, weight=1)
        self.grid_columnconfigure(0, weight=1)

        self.all_products = []
        self.filtered_products = []
        self.category_cache = []
        
        # Pagination
        self.page_size = 20  # Increased from 15 for better performance
        self.current_page = 1
        self.total_pages = 1
        
        # Performance flags
        self.is_loading = False
        self.filter_timer = None
        self.last_search = ""
        self.last_category = ""

        self.create_widgets()
        self.load_products_async()
    
    def create_widgets(self):
        """Create UI components - optimized"""
        # Header
        header = ctk.CTkFrame(self, fg_color="transparent")
        header.grid(row=0, column=0, sticky="ew", padx=20, pady=20)
        header.grid_columnconfigure(2, weight=1)
        
        back_btn = ctk.CTkButton(
            header, text="← Back", width=80, height=40,
            command=self.go_back, fg_color="#666666", hover_color="#555555",
            font=("Helvetica", 12, "bold")
        )
        back_btn.grid(row=0, column=0, padx=(0, 10))
        
        title = ctk.CTkLabel(
            header, text="📊 Inventory Management",
            font=("Helvetica", 28, "bold"), text_color="#1f1f1f"
        )
        title.grid(row=0, column=1, sticky="w", padx=(0, 10))
        
        # Right side buttons
        button_group = ctk.CTkFrame(header, fg_color="transparent")
        button_group.grid(row=0, column=3, sticky="e")
        
        add_btn = ctk.CTkButton(
            button_group, text="+ Add Product", width=120, height=40,
            command=self.open_add_dialog, fg_color="#059669", hover_color="#047857",
            font=("Helvetica", 12, "bold")
        )
        add_btn.pack(side="left", padx=5)
        
        import_btn = ctk.CTkButton(
            button_group, text="📥 Import Excel", width=120, height=40,
            command=self.import_excel, fg_color="#7C3AED", hover_color="#6D28D9",
            font=("Helvetica", 12, "bold")
        )
        import_btn.pack(side="left", padx=5)
        
        refresh_btn = ctk.CTkButton(
            button_group, text="🔄 Refresh", width=100, height=40,
            command=self.load_products_async, fg_color="#2563EB", hover_color="#1E40AF",
            font=("Helvetica", 12, "bold")
        )
        refresh_btn.pack(side="left", padx=5)
        
        # Search bar
        search_frame = ctk.CTkFrame(self, fg_color="transparent")
        search_frame.grid(row=1, column=0, sticky="ew", padx=20, pady=(0, 15))
        
        search_label = ctk.CTkLabel(search_frame, text="🔍 Search by Product Name:", font=("Helvetica", 12, "bold"), text_color="#1f1f1f")
        search_label.pack(side="left", padx=(0, 10))
        
        self.search_entry = ctk.CTkEntry(search_frame, placeholder_text="Type product name...", width=300, height=35, font=("Helvetica", 11))
        self.search_entry.pack(side="left", padx=5)
        self.search_entry.bind("<KeyRelease>", self.on_search_change)
        
        # Category filter
        category_label = ctk.CTkLabel(search_frame, text="📁 Filter by Category:", font=("Helvetica", 12, "bold"), text_color="#1f1f1f")
        category_label.pack(side="left", padx=(20, 10))
        
        self.category_combo = ctk.CTkComboBox(
            search_frame, state="readonly", width=200, height=35, font=("Helvetica", 11),
            values=["All Categories"],
            command=lambda v: self.schedule_filter()
        )
        self.category_combo.pack(side="left", padx=5)
        self.category_combo.set("All Categories")
        
        # Table frame with fixed column layout
        table_frame = ctk.CTkFrame(self, fg_color="white", corner_radius=10)
        table_frame.grid(row=2, column=0, sticky="nsew", padx=20, pady=(0, 20))
        table_frame.grid_rowconfigure(1, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)
        
        # Column width configuration (proportional)
        self.col_widths = {
            'name': 0.25,      # 25%
            'category': 0.15,  # 15%
            'qty': 0.12,       # 12%
            'pp': 0.15,        # 15%
            'sp': 0.15,        # 15%
            'actions': 0.18    # 18%
        }
        
        # Table header - fixed width
        header_frame = ctk.CTkFrame(table_frame, fg_color="#f0f0f0", corner_radius=8)
        header_frame.grid(row=0, column=0, sticky="ew", padx=5, pady=5)
        header_frame.grid_columnconfigure(0, weight=int(self.col_widths['name']*100))
        header_frame.grid_columnconfigure(1, weight=int(self.col_widths['category']*100))
        header_frame.grid_columnconfigure(2, weight=int(self.col_widths['qty']*100))
        header_frame.grid_columnconfigure(3, weight=int(self.col_widths['pp']*100))
        header_frame.grid_columnconfigure(4, weight=int(self.col_widths['sp']*100))
        header_frame.grid_columnconfigure(5, weight=int(self.col_widths['actions']*100))
        
        headers = ["Product Name", "Category", "Quantity", "Purchase Price", "Selling Price", "Actions"]
        header_configs = [
            ("w", "#1f1f1f"),  # name - left align
            ("w", "#1f1f1f"),  # category - left align
            ("center", "#1f1f1f"),  # qty - center align
            ("center", "#1f1f1f"),  # pp - center align
            ("center", "#1f1f1f"),  # sp - center align
            ("center", "#1f1f1f")   # actions - center align
        ]
        
        for i, (h, (anchor, color)) in enumerate(zip(headers, header_configs)):
            label = ctk.CTkLabel(
                header_frame, text=h, font=("Helvetica", 12, "bold"),
                text_color=color, anchor=anchor
            )
            label.grid(row=0, column=i, padx=12, pady=10, sticky="ew")
        
        # Scrollable frame
        self.scrollable_frame = ctk.CTkScrollableFrame(
            table_frame, fg_color="transparent"
        )
        self.scrollable_frame.grid(row=1, column=0, sticky="nsew", padx=5, pady=5)
        self.scrollable_frame.grid_columnconfigure(0, weight=1)
        self.scrollable_frame._scrollbar.configure(width=12)
        
        self.table_frame_ref = table_frame
    
    def load_products_async(self):
        """Load products in background thread"""
        if self.is_loading:
            return
        
        self.is_loading = True
        threading.Thread(target=self._load_products_thread, daemon=True).start()
    
    def _load_products_thread(self):
        """Background thread for loading products - optimized"""
        try:
            # Fetch data once
            self.all_products = get_all_products()
            
            # Extract categories in single pass
            categories = sorted(set(p[2] for p in self.all_products if p[2]))
            self.category_cache = categories
            
            # Update UI in main thread
            self.after(0, self._update_ui_after_load)
        except Exception as e:
            print(f"✗ Error loading products: {e}")
            self.is_loading = False
    
    def _update_ui_after_load(self):
        """Update UI after loading products"""
        self.category_combo.configure(values=["All Categories"] + self.category_cache)
        self.category_combo.set("All Categories")
        self.last_search = ""
        self.last_category = ""
        self.schedule_filter()
        self.is_loading = False
    
    def on_search_change(self, event):
        """Handle search input with debouncing"""
        self.schedule_filter()
    
    def schedule_filter(self):
        """Schedule filter with debouncing"""
        if self.filter_timer:
            self.after_cancel(self.filter_timer)
        self.filter_timer = self.after(100, self.filter_products)

    def filter_products(self):
        """Filter products - optimized with early returns"""
        if self.is_loading:
            return
        
        search_text = self.search_entry.get().lower()
        selected_category = self.category_combo.get()
        
        # Skip filtering if nothing changed
        if search_text == self.last_search and selected_category == self.last_category:
            return
        
        self.last_search = search_text
        self.last_category = selected_category

        filtered = self.all_products

        # Optimized filtering in single pass
        if search_text or selected_category != "All Categories":
            filtered = [
                p for p in filtered
                if (not search_text or search_text in p[1].lower()) and
                   (selected_category == "All Categories" or p[2] == selected_category)
            ]

        self.filtered_products = filtered
        self.total_pages = max(1, (len(filtered) + self.page_size - 1) // self.page_size)
        self.current_page = 1
        self.display_page()

    def display_page(self):
        """Display current page - lazy rendering"""
        # Clear only visible widgets
        for widget in self.scrollable_frame.winfo_children():
            widget.destroy()

        start_idx = (self.current_page - 1) * self.page_size
        end_idx = start_idx + self.page_size
        page_products = self.filtered_products[start_idx:end_idx]

        if not page_products:
            empty_label = ctk.CTkLabel(
                self.scrollable_frame,
                text="No products found." if self.filtered_products else "No products added yet.",
                text_color="#999999", font=("Helvetica", 12, "bold")
            )
            empty_label.pack(pady=30)
        else:
            # Create all rows efficiently
            for product in page_products:
                self._create_product_row(product)

        self._update_pagination()

    def _create_product_row(self, product):
        """Create a single product row with fixed column widths"""
        prod_id, name, category, qty, pp, sp = product

        row_frame = ctk.CTkFrame(
            self.scrollable_frame, fg_color="white", corner_radius=8, border_width=1, border_color="#e0e0e0"
        )
        row_frame.pack(fill="x", padx=5, pady=4)
        
        # Set exact column proportions matching header
        row_frame.grid_columnconfigure(0, weight=int(self.col_widths['name']*100))
        row_frame.grid_columnconfigure(1, weight=int(self.col_widths['category']*100))
        row_frame.grid_columnconfigure(2, weight=int(self.col_widths['qty']*100))
        row_frame.grid_columnconfigure(3, weight=int(self.col_widths['pp']*100))
        row_frame.grid_columnconfigure(4, weight=int(self.col_widths['sp']*100))
        row_frame.grid_columnconfigure(5, weight=int(self.col_widths['actions']*100))

        # Color coding for low stock
        qty_color = "#DC2626" if qty <= 5 else "#1f1f1f"
        qty_text = f"{qty} {'⚠️' if qty <= 5 else ''}"

        # Column data with alignment
        col_data = [
            (name, 0, "w", "#1f1f1f", True),           # name - left
            (category or "General", 1, "w", "#666666", True),  # category - left
            (qty_text, 2, "center", qty_color, True),  # qty - center
            (f"₨{pp:.2f}", 3, "center", "#1f1f1f", True),     # pp - center
            (f"₨{sp:.2f}", 4, "center", "#1f1f1f", True),     # sp - center
        ]

        for text, col, anchor, color, bold in col_data:
            label = ctk.CTkLabel(
                row_frame, text=text, font=("Helvetica", 11, "bold" if bold else "normal"),
                anchor=anchor, text_color=color
            )
            label.grid(row=0, column=col, padx=12, pady=10, sticky="ew")

        # Action buttons frame
        action_frame = ctk.CTkFrame(row_frame, fg_color="transparent")
        action_frame.grid(row=0, column=5, padx=8, pady=10, sticky="ew")
        action_frame.grid_columnconfigure(0, weight=1)
        action_frame.grid_columnconfigure(1, weight=1)

        edit_btn = ctk.CTkButton(
            action_frame, text="✎", width=35, height=35,
            font=("Helvetica", 14, "bold"), fg_color="#2563EB", hover_color="#1E40AF",
            command=lambda pid=prod_id: self.open_edit_dialog(pid)
        )
        edit_btn.grid(row=0, column=0, padx=2)

        del_btn = ctk.CTkButton(
            action_frame, text="✕", width=35, height=35,
            font=("Helvetica", 14, "bold"), fg_color="#DC2626", hover_color="#B91C1C",
            command=lambda pid=prod_id: self.delete_product(pid)
        )
        del_btn.grid(row=0, column=1, padx=2)

    def _update_pagination(self):
        """Update pagination controls"""
        if hasattr(self, 'pagination_frame'):
            self.pagination_frame.destroy()

        self.pagination_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.pagination_frame.grid(row=3, column=0, pady=(0, 20))

        prev_btn = ctk.CTkButton(
            self.pagination_frame, text="← Previous", width=120,
            command=self.prev_page, fg_color="#2563EB"
        )
        prev_btn.pack(side="left", padx=5)
        
        page_label = ctk.CTkLabel(
            self.pagination_frame,
            text=f"Page {self.current_page} of {self.total_pages}",
            font=("Helvetica", 12, "bold")
        )
        page_label.pack(side="left", padx=10)
        
        next_btn = ctk.CTkButton(
            self.pagination_frame, text="Next →", width=120,
            command=self.next_page, fg_color="#2563EB"
        )
        next_btn.pack(side="left", padx=5)

    def next_page(self):
        """Go to next page"""
        if self.current_page < self.total_pages:
            self.current_page += 1
            self.display_page()

    def prev_page(self):
        """Go to previous page"""
        if self.current_page > 1:
            self.current_page -= 1
            self.display_page()
    
    def import_excel(self):
        """Import products from Excel file"""
        file_path = filedialog.askopenfilename(
            title="Select Excel File to Import",
            filetypes=[("Excel Files", "*.xlsx *.xls"), ("All Files", "*.*")]
        )
        
        if not file_path:
            return
        
        threading.Thread(
            target=self._import_excel_thread,
            args=(file_path,),
            daemon=True
        ).start()
    
    def _import_excel_thread(self, file_path):
        """Background thread for importing Excel"""
        try:
            workbook = load_workbook(file_path)
            sheet = workbook.active
            
            headers = [cell.value for cell in sheet[1]]
            required_headers = ["Product Name", "Category", "Quantity", "Purchase Price", "Selling Price"]
            
            if not all(h in headers for h in required_headers):
                error_msg = (
                    f"⚠️  Invalid Excel format!\n\n"
                    f"Required columns:\n"
                    f"• Product Name\n"
                    f"• Category\n"
                    f"• Quantity\n"
                    f"• Purchase Price\n"
                    f"• Selling Price"
                )
                self.after(0, lambda: self.show_message(error_msg, "Format Error"))
                return
            
            col_indices = {h: idx + 1 for idx, h in enumerate(headers) if h in required_headers}
            
            imported_count = 0
            skipped_count = 0
            errors = []
            
            # Batch process rows
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
                try:
                    product_name = row[col_indices["Product Name"] - 1].value
                    category = row[col_indices["Category"] - 1].value
                    quantity = row[col_indices["Quantity"] - 1].value
                    purchase_price = row[col_indices["Purchase Price"] - 1].value
                    selling_price = row[col_indices["Selling Price"] - 1].value
                    
                    if not product_name:
                        skipped_count += 1
                        continue
                    
                    try:
                        quantity = int(quantity) if quantity else 0
                        purchase_price = float(purchase_price) if purchase_price else 0
                        selling_price = float(selling_price) if selling_price else 0
                    except (ValueError, TypeError):
                        errors.append(f"Row {row_idx}: Invalid data type")
                        continue
                    
                    success, msg = add_product(
                        str(product_name).strip(),
                        str(category).strip() if category else "General",
                        quantity,
                        purchase_price,
                        selling_price
                    )
                    
                    if success:
                        imported_count += 1
                    else:
                        skipped_count += 1
                
                except Exception as e:
                    errors.append(f"Row {row_idx}: {str(e)}")
            
            summary = f"✓ Import Complete!\n\nImported: {imported_count}\nSkipped: {skipped_count}"
            if errors:
                summary += f"\nErrors: {len(errors)}"
            
            self.after(0, lambda: self.show_message(summary, "Import Summary"))
            self.after(0, self.load_products_async)
            
        except PermissionError:
            self.after(0, lambda: self.show_message(
                "⚠️  Permission Denied!\n\nClose the file in Excel first.",
                "Error"
            ))
        except Exception as e:
            self.after(0, lambda: self.show_message(f"⚠️  Error:\n\n{str(e)}", "Error"))
    
    def open_add_dialog(self):
        """Open dialog to add product"""
        dialog = ctk.CTkInputDialog(text="Product Name:", title="Add Product")
        name = dialog.get_input()
        if not name:
            return
        
        category = self.show_input("Category:", name)
        if not category:
            return
        
        qty_str = self.show_input("Quantity:", category)
        if not qty_str or not qty_str.isdigit():
            self.show_message("Invalid quantity", "⚠")
            return
        
        pp_str = self.show_input("Purchase Price:", qty_str)
        if not pp_str:
            return
        
        sp_str = self.show_input("Selling Price:", pp_str)
        if not sp_str:
            return
        
        try:
            qty = int(qty_str)
            pp = float(pp_str)
            sp = float(sp_str)
            
            success, msg = add_product(name, category, qty, pp, sp)
            if success:
                self.show_message(msg, "✓")
                self.load_products_async()
            else:
                self.show_message(msg, "⚠")
        except ValueError:
            self.show_message("Invalid input values", "⚠")
    
    def open_edit_dialog(self, product_id):
        """Open dialog to edit product"""
        product = next((p for p in self.all_products if p[0] == product_id), None)
        
        if not product:
            return
        
        prod_id, name, category, qty, pp, sp = product
        
        dialog = ctk.CTkInputDialog(text=f"Product Name ({name}):", title="Edit Product")
        new_name = dialog.get_input()
        if new_name is None:
            return
        new_name = new_name or name
        
        category = self.show_input(f"Category ({category}):", name)
        category = category or product[2]
        
        qty_str = self.show_input(f"Quantity ({qty}):", category)
        qty = int(qty_str) if qty_str and qty_str.isdigit() else product[3]
        
        pp_str = self.show_input(f"Purchase Price ({pp}):", str(qty))
        pp = float(pp_str) if pp_str else product[4]
        
        sp_str = self.show_input(f"Selling Price ({sp}):", str(pp))
        sp = float(sp_str) if sp_str else product[5]
        
        success, msg = update_product(prod_id, new_name, category, qty, pp, sp)
        if success:
            self.show_message(msg, "✓")
            self.load_products_async()
        else:
            self.show_message(msg, "⚠")
    
    def delete_product(self, product_id):
        """Delete a product"""
        success, msg = delete_product(product_id)
        if success:
            self.show_message(msg, "✓")
            self.load_products_async()
        else:
            self.show_message(msg, "⚠")
    
    def show_input(self, prompt, title):
        """Show input dialog"""
        dialog = ctk.CTkInputDialog(text=prompt, title=title)
        return dialog.get_input()
    
    def show_message(self, msg, icon):
        """Show message dialog"""
        dialog = ctk.CTkInputDialog(text=msg, title=icon)
        dialog.get_input()
